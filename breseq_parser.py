#! /usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook, styles
from collections import OrderedDict
from bs4 import BeautifulSoup
import pandas
from unidecode import unidecode
import pathlib
from typing import *
import argparse
import os
from pprint import pprint
from functools import partial

print = partial(print, flush = True)
Table = List[Dict[str, str]]
# TODO expand user folder with ~ for the -d flag
DEBUG = os.name == 'nt'

if not DEBUG:
	parser = argparse.ArgumentParser(
		description = "This is a Breseq Mutation Parser.  It Currently only SNPs, Missing Coverage, "
					  "and New Junction Evidence (wont output junction repeats).  "
					  "In order to run this program please put all the Breseq directores into a master directory and then "
					  "parse that directory with the program's -d flag.")

	parser.add_argument(
		'-d', '--directory',
		action = "store",
		help = "Use this flag to indicate the index file you would like to parse",
		dest = "directory"
	)
	parser.add_argument(
		'-f', '--format',
		action = "store",
		help = "format of the output file.",
		dest = 'filetype',
		choices = ['csv', 'tsv', 'xlsx'],
		default = 'xlsx'
	)
	parser.add_argument(
		'-o', '--output',
		action = "store",
		help = "Name of the output file(s) or folder. Defaults to './breseq_output'",
		default = 'breseq_output',
		dest = 'filename'
	)

	args = parser.parse_args()

else:

	class Parser:
		directory: str
		filetype: str
		prefix: str

		def __init__(self, a, b, c):
			self.directory = a
			self.filetype = b
			self.filename = c


	test_folder = pathlib.Path(__file__).parent / 'test_data'
	args = Parser(test_folder, 'xlsx', test_folder.with_name('test_output.xlsx'))


def toNumber(string: str) -> int:
	""" Converts a string to a number"""
	try:
		string = int(string.replace(',', ''))
	except ValueError:
		pass
	return string


class Breseq:
	def __init__(self, options):

		self.options = options
		self.data_folder = pathlib.Path(self.options.directory)
		self.snp_table = list()
		self.coverage_table = list()
		self.junction_table = list()
		for folder in self.data_folder.iterdir():
			if not folder.is_dir(): continue
			print("parsing ", folder)
			snp_table, coverage_table, junction_table = self.parseOutputFolder(folder)

			self.snp_table += snp_table
			self.coverage_table += coverage_table
			self.junction_table += junction_table

		self.snp_table = pandas.DataFrame(self.snp_table)
		self.coverage_table = pandas.DataFrame(self.coverage_table)
		self.junction_table = pandas.DataFrame(self.junction_table)
		self.generateComparisonTable(self.snp_table)

	def parseOutputFolder(self, folder: pathlib.Path) -> Tuple[Table, Table, Table]:

		index_file = folder / "output" / "index.html"
		print("\tIndex File: ", index_file)
		if not index_file.exists():
			print("\tThe index.html file is missing. Ignoring folder.")
			return [], [], []
		sample_name = folder.name
		snp_headers, snp_table, coverage_soup, junction_soup = self._parseIndexFile(index_file)
		parsed_snp_table = self._parsePredictedMutations(sample_name, snp_headers, snp_table)
		coverage_table = self._parseCoverage(sample_name, coverage_soup)
		junction_table = self._parseJunctions(sample_name, junction_soup)
		return parsed_snp_table, coverage_table, junction_table

	@staticmethod
	def _extractIndexFileTables(soup: BeautifulSoup) -> Tuple[List[str], BeautifulSoup, BeautifulSoup]:
		alph_soup = str(soup)
		# print(alph_soup)
		begin_snp_header_string = r'<th>evidence</th>'
		# begin_snp_header_string = r'Predicted mutations</th></tr><tr>'
		end_snp_header_string = '<!-- Item Lines -->'

		begin_snp_header = alph_soup.find(begin_snp_header_string)
		end_snp_header = alph_soup.find(end_snp_header_string)

		snp_header_full = alph_soup[begin_snp_header:end_snp_header]
		snp_header = snp_header_full[end_snp_header:]

		snp_header_soup = BeautifulSoup(snp_header_full, 'lxml')
		snp_header_soup = [i.text for i in snp_header_soup.find_all('th')]

		if DEBUG and False:
			print("Begin snp header index: ", begin_snp_header)
			print("End snp header index: ", end_snp_header)
			print()
			print("_extractIndexFilesTables")
			print("Full Header: ", type(snp_header_full))
			print(snp_header_full)
			print()
			print("Short Header:", type(snp_header))
			print(len(snp_header))
			print(snp_header)
			print()
			print("BS4 Header:")
			print(snp_header_soup)

		begin_umc = alph_soup.find(
			'<tr><th align="left" class="missing_coverage_header_row" colspan="11">Unassigned missing coverage evidence</th></tr>')
		end_umc = alph_soup.find(
			'<th align="left" class="new_junction_header_row" colspan="12">Unassigned new junction evidence</th>')
		coverage_string = alph_soup[begin_umc:end_umc]
		junction_string = alph_soup[end_umc:]
		coverage_soup = BeautifulSoup(coverage_string, 'lxml')
		junction_soup = BeautifulSoup(junction_string, 'lxml')
		return snp_header_soup, coverage_soup, junction_soup

	def _parseIndexFile(self, filename: pathlib.Path):
		with open(filename, 'r') as file1:
			contents = file1.read()

		soup = BeautifulSoup(contents, 'lxml')

		normal_table = soup.find_all(attrs = {'class': 'normal_table_row'})
		poly_table = soup.find_all(attrs = {'class': 'polymorphism_table_row'})
		if len(normal_table):
			snp_table = normal_table
		else:
			snp_table = poly_table
		snp_table = normal_table + poly_table
		snp_header_soup, coverage_soup, junction_soup = self._extractIndexFileTables(soup)

		return snp_header_soup, snp_table, coverage_soup, junction_soup

	@staticmethod
	def _parsePredictedMutations(sample_name: str, headers: List[str], rows: List) -> Table:

		converted_table = list()

		for tag in rows:
			values = [v.text for v in tag.find_all('td')]

			if len(values) > 1:
				row = {k: v for k, v in zip(headers, values)}
				row['Sample'] = sample_name

				row['position'] = toNumber(row['position'])
				try:
					row['freq %'] = float(row['freq'][:-1])
					row.pop('freq')
				except KeyError:
					pass
				converted_table.append(row)
		return converted_table

	@staticmethod
	def _parseCoverage(sample_name: str, coverage: BeautifulSoup) -> Table:
		coverage_table = list()
		rows = coverage.find_all('tr')
		if len(rows) == 0:
			print("\tCould not parse the coverage table.")
			return coverage_table
		column_names = [i.text for i in rows[1].find_all('th')]

		for index, tag in enumerate(rows[2:]):
			values = tag.find_all('td')

			if len(values) > 1:
				row = [('Sample', sample_name)] + [(k, v.get_text()) for k, v in zip(column_names, values)]
				row = OrderedDict(row)

				row['start'] = toNumber(row['start'])
				row['end'] = toNumber(row['end'])
				row['size'] = toNumber(row['size'])
				coverage_table.append(row)

		return coverage_table

	@staticmethod
	def _parseJunctions(sample_name: str, junctions: BeautifulSoup) -> Table:
		rows = junctions.find_all('tr')
		if len(rows) == 0:
			print("\tCould not parse Junctino table.")
			return list()
		column_names_a = ['0', '1'] + [unidecode(i.get_text()) for i in rows.pop(0).find_all('th')][1:]

		column_names_a[4] = '{} ({})'.format(column_names_a[4], 'single')
		column_names_b = [i for i in column_names_a if i not in ['reads (cov)', 'score', 'skew', 'freq', '0']]
		junction_table = list()
		for a_row, b_row in zip(rows[::2], rows[1::2]):
			a_values = [unidecode(i.get_text()) for i in a_row.find_all('td')]
			b_values = [unidecode(i.get_text()) for i in b_row.find_all('td')]

			a_row = {unidecode(k): v for k, v in zip(column_names_a, a_values)}
			b_row = {unidecode(k): v for k, v in zip(column_names_b, b_values)}
			a_row['Sample'] = sample_name
			b_row['Sample'] = sample_name
			junction_table.append(a_row)
			junction_table.append(b_row)
		return junction_table

	@staticmethod
	def generateComparisonTable(snp_table: pandas.DataFrame) -> Optional[pandas.DataFrame]:
		# Sample	annotation	description	evidence	gene	mutation	position	seq id
		try:
			all_samples = set(snp_table['Sample'].values)
			columns = [unidecode(i) for i in snp_table.columns.values]

			sequence_group = snp_table.groupby(by = ['seq id', 'position'])
			comparison = list()
			for element, sequence in sequence_group:
				seq_id, position = element
				samples = sequence['Sample'].values

				char = 'X' if len(samples) == 1 else '.'

				comparison_row = {k: char for k in samples}
				comparison_row['seq id'] = seq_id
				comparison_row['position'] = position
				comparison_row['all'] = '.' if len(samples) == len(all_samples) else ''
				comparison.append(comparison_row)

			comparison_table = pandas.DataFrame(comparison)
		except:
			comparison_table = None
		return comparison_table

	def _formatComparisonWorksheet(self, worksheet):

		for character in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
			for index in range(1, len(self.snp_table)):
				cell_index = character + str(index)
				cell = worksheet[cell_index]
				print(cell)
				value = cell.value
				if value == '.':

					cell.fill = styles.Fill(bgColor = styles.Color('00000000'), fill_type = 'solid')
				elif value == 'X':

					cell.fill = styles.PatternFill(bgColor = styles.Color('12345678'), fill_type = 'solid')

		return worksheet

	def save(self, filename = None, filetype = None):
		if filetype is None:
			filetype = args.filetype
		filetype = filetype.lower()

		if filename is None:
			filename = args.filename
		filename = pathlib.Path(filename)
		if filename.is_dir():
			filename = filename / 'breseq_output'

		print("Saving to ", filename)

		if filetype == 'xlsx':
			self.to_excel(filename)
		else:
			self.to_csv(filename, filetype)

	def to_excel(self, filename):
		if isinstance(filename, str):
			filename = pathlib.Path(filename)
		filename = filename.with_suffix('.xlsx')
		writer = pandas.ExcelWriter(filename)
		include_index = False

		self.snp_table.to_excel(writer, 'snps', index = include_index)

		self.coverage_table.to_excel(writer, 'coverage', index = include_index)

		self.junction_table.to_excel(writer, 'junctions', index = include_index)
		try:
			comparison_table = self.generateComparisonTable(self.snp_table)
		except KeyError:
			comparison_table = None
		if comparison_table is not None:
			comparison_table.to_excel(writer, 'snp comparison', index = include_index)

		writer.close()

		wb = load_workbook(filename)
		# cws = wb.get_sheet_by_name('snp comparison')
		# cws = self._formatComparisonWorksheet(cws)
		ws = wb.get_sheet_by_name('junctions')

		merge_columns = [
			c
			for c in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
			if ws[c + '1'].value in ['Sample', 0, '0', 'freq', 'product', 'score']
		]

		for x in range(0, len(self.junction_table), 2):
			for column in merge_columns:
				cells_to_merge = '{0}{1}:{0}{2}'.format(column, x + 2, x + 3)
				ws.merge_cells(cells_to_merge)

		wb.save(filename)

	def to_csv(self, folder: Union[str, pathlib.Path], filetype):
		if isinstance(folder, str):
			folder = pathlib.Path(folder)
		extension = 'tsv' if filetype == 'tsv' else 'csv'
		snp_filename = str(folder.with_name(folder.stem + '.snp.' + extension).absolute())
		coverage_filename = str(folder.with_name(folder.stem + '.coverage.' + extension).absolute())
		junction_filename = str(folder.with_name(folder.stem + '.junction.' + extension).absolute())

		delimiter = '\t' if filetype == 'tsv' else ','
		include_index = False

		self.snp_table.to_csv(snp_filename, sep = delimiter, index = include_index)
		self.coverage_table.to_csv(coverage_filename, sep = delimiter, index = include_index)
		self.junction_table.to_csv(junction_filename, sep = delimiter, index = include_index)

	def to_vdf(self):
		raise NotImplementedError

if __name__ == "__main__":
	data_folder = args.directory
	output_file = pathlib.Path(args.filename)
	if not data_folder or not pathlib.Path(data_folder).is_dir():
		print("This is not a valid folder: ", data_folder)
		print("Please Enter a valid Directory to parse, try the --help flag if you have questions, exiting!")
		exit(1)

	obj = Breseq(args)
	obj.save(args.filename, args.filetype)
